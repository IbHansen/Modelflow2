# -*- coding: utf-8 -*-
"""
Created on Fri Feb 12 07:04:02 2016

@author: ibh

Takes all formula's  from a excel work book and translates each to the equivalent expression. 
Openpyxl is the fastest library but it can not deal all values. Therefor xlwings is also used. 
But only to read repeated formula's which inly will show as '='



also defines function used whenusing xlwings to automate excel 

"""
import pandas as pd 
import networkx as nx 
import openpyxl
from openpyxl import load_workbook
from openpyxl.formula import Tokenizer
from openpyxl.utils import get_column_letter
from openpyxl.utils import cols_from_range,rows_from_range
# try:
#     import xlwings as xw
# except:
#     ...
# import networkx as nx

import matplotlib.pylab  as plt 
import seaborn as sns
import os 
import sys
from pathlib import Path

import modelclass as mc

DEBUG = 0

def findequations(name):
    '''Takes all formula's  from a excel work book and translates each to the equivalent expression. 
    
        Multicell ranges are expanded to a comma separated list. \n
        The ordinary operators and the SUM function can be handled. If you need more functions. You have to impelent them in the modelclass. 
        
        In the model each cell reference is prefixed by <sheet name>_
        
        Openpyxl is the fastest library and it has a tokenizer
        but it can not read all values. 
        
        Therefor xlwings is used to read repeated formula's which Openpyxl will show as '='

    input:
        :name: Location of a excel sheeet
    
         
        Returns:
             :modeldic: A dictionary with formulars keyed by cell reference   
         
    '''
    outdic={}
    wb = load_workbook(name, read_only=True,data_only=False) # to read the spresdsheet first save as xml then write it again 
    try:
        wb2 = xw.Book(name)    # the same worksheet in xlwings
    except:
        ... 
    # breakpoint()
    
    allsheets = wb.sheetnames
    for wsname in allsheets:
        ws=wb[wsname]
        try:
            ws2=wb2.sheets(wsname)   # the same sheet but in xlwings
        except: 
            ...
            
        formulacell = [c for row in ws.rows for c in row if c.value != None and c.data_type == 'f'] 
        for cell in formulacell:
            cellref=get_column_letter(cell.column)+str(cell.row)
            if DEBUG : print('This cell:',cellref,cell.data_type,cell.value)
            if cell.value == '=' :
                print(f'Repeat cell = so xlwings has to be imported {cell=}')
                frml = cell.value if cell.value != '=' else ws2.range(cellref).formula # To avoid emty repeating formula'rs 
            else:
                frml = cell.value 
            tok=Tokenizer(frml)
            if DEBUG and False : print("\n".join("%19s%15s%9s" % (t.value, t.type, t.subtype) for t in tok.items))
            # left hand term is <worksheet>!<column><row>= 
            lhs=wstrans(wsname) + get_column_letter(cell.column)+str(cell.row)
            out=[lhs+'=']
            for t in tok.items:            
                if t.subtype == "RANGE":
                    #Find or create the sheetname  
                    sheet0 = t.value.split('!')[0] if '!' in t.value else wsname
                    sheet  =  wstrans(sheet0)
                    # print(t.value,'---->')
                    # Get all the cells in the range columwize 
                    # the nested list comprehension makes the list works for square ranges. 
                    # the split construct drops the sheet name from the range name if any 
                    thisrange=[sheet+i for subtupler in cols_from_range((t.value.split('!')[-1])) for i in subtupler]

                    # put a ',' between each element i the list 
                    thistext=','.join(thisrange)
                    #print(thisrange)
                    out=out + [thistext]
                else:
                    out.append(t.value)
            #create the equation and get rid of the !
            equation=''.join(out).replace('!','_')
            outdic[lhs]=equation
            #print(equation)
    try:        
        wb2.close()
    except: 
        ... 
    return outdic   
    
def showcells(name):
    '''Finds values in a excel workbook with a value different from 0 
    '''
    wb = load_workbook(name, read_only=True,data_only=False) # to read the spresdsheet first save as xml then write it again 
    allsheets = wb.sheetnames
    for wsname in allsheets:
        ws=wb[wsname]
        for row in ws.rows:
            for c in row:
                if c.value != None:
                    print(wsname,get_column_letter(c.column ),c.row,c.data_type,c.value)


def findvalues(name):
    '''Finds numerical values in a excel workbook with a value different from 0 
    '''
    wb = load_workbook(name, read_only=True,data_only=True) # to read the spresdsheet first save as xml then write it again 
    allsheets = wb.sheetnames
    values=[]
    for wsname in allsheets:
        ws=wb[wsname]
        twsname = wstrans(wsname)
        values+=[(twsname+get_column_letter(c.column )+str(c.row),c.value) 
           for row in ws.rows for c in row 
             if c.value != None and c.data_type == 'n' ]  
    return values 
    
def wstrans(wsname):
    res = '_'+wsname.replace("'","").replace(' - ','_').replace(' ','_').replace('-','_')+'_'
    return res.upper() 

def findcoordinates(name):
    '''Finds the cell references matching the codes in a LCR workbook from EBA 
    
    This is needed for the mapping of the raw data to the excel cell refereces. 
    
    input:
        :name: Location of a excel sheeet
    
    returns:
        :coldf: Dataframe with mapping between excel column and EBA columns_code
        :rowdf: Dataframe with row with mapping between excel row and EBS data row_code
    '''
    wb = load_workbook(name, read_only=True,data_only=True) # to read the spresdsheet first save as xml then write it again 
    allsheets = wb.sheetnames
    
    colcodes=[]
    rowcodes=[]
    
    for wsname in allsheets:
        ws=wb[wsname]
        try:
        #find the anchor for the row and columns id, the first 
            cell = [c for row in ws.rows for c in row if c.value and c.data_type == 's' and 'Row'== c.value][0]  
            # find the numeric values in the column below the anchor, only the digits and remenber python index starts with 0 while excels index starts with 1 
     
            rowcodes += [(wsname , c.value , c.row ) for r in ws.rows for c in r  
                        if c.value and c.column == cell.column and c.row > cell.row and 
                              ( c.data_type == 'n' or (c.data_type == 's' and c.value.isdigit()) )]
            
    #         finds the numeric values in the row at the right of the anchor. 
            # c.rows returns a generator (probably because te potential for a huge number) therefor the list(c.rows)
            colcodes += [(wsname , c.value , get_column_letter(c.column ) )  for c in list(ws.rows)[cell.row-1] 
                        if c.value and c.column > cell.column and (c.data_type == 'n' or (c.data_type == 's' and c.value.isdigit()))] 
        except:  # Ok this ws did not have an ancor cell 
            pass
    coldf = pd.DataFrame(colcodes,columns=['sheet','colcode','col'])
    rowdf = pd.DataFrame(rowcodes,columns=['sheet','rowcode','row'])
    return coldf,rowdf

def getexcelmodel(name):
    ''' Creates a model instance from a excel sheet
    SUM is replaced by SUM_EXCEL which is a function in the modelclass 
    
    In the excel formulars this function accepts ordinary operators and SUM in excel sheets 
    
    input:
        :name: Location of a excel sheeet
    
    returns:
        :model: A model instance with the formulars of the excel sheet
        :para: A list of values in the sheet which matches exogeneous variables in the model 

    '''
    modelname = Path(name).stem
    eqdic   = findequations(name)
    eqdic2  = {i : eq.replace('SUM(','SUM_EXCEL(')  for i,eq in eqdic.items()}                
    fdic    = {i : 'Frml xx '+eqdic2[i] + r' $ ' for i,eq in eqdic2.items()}                
    f       = '\n'.join([eq for i,eq in fdic.items()])                
    _mmodel=mc.model(f,modelname=modelname)               
    zz = findvalues(name)
    para = [z for z in zz if z[0] in _mmodel.exogene]   # find all values which match a exogeneous variable in model
    return _mmodel,para

#%% now functions related to running xlsheets 


def indextrans(index):
    '''
    Transforms a period index to excel acceptable datatype 
    
    
    '''    
    out = [i.year if type(index) == pd.core.indexes.period.PeriodIndex 
                     else int(i) for i in index]
    return out 

def df_to_sheet(name,df,wb,after=None):
    try:
        wb.sheets[name].delete()
    except:
        pass  
    
    try:
        sht = wb.sheets.add(name,after=after)
    except Exception as e :
        print('no sheet added',str(e))
    df_ = df.copy()
    df_.index = indextrans(df.index)
    sht.range('A1').value = df_.T
    active_window = wb.app.api.ActiveWindow
    active_window.FreezePanes = False
    active_window.SplitColumn = 1
    active_window.SplitRow = 1
    active_window.FreezePanes = True
    sht.autofit(axis="columns")
    sht[(2,25)].select()
    return sht

def obj_to_sheet(name,obj,wb,after=None):
    # breakpoint()
    try:
        wb.sheets[name].delete()
    except:
        pass
    
    try:
        sht = wb.sheets.add(name,after=after)
    except Exception as e:
        print(str(e))
        print('no sheet added ')
    sht.range('A1').value=obj
    

def sheet_to_df(wb,name):
    df = wb.sheets[name].range('A1').options(pd.DataFrame, expand='table').value.T
    df.index = indextrans(df.index)
    return df 

def sheet_to_dict(wb,name,integers=None):
    ''' transform the named sheet to a python dict. If we need a integer it has to be in the integer set'''
    
    integers_ = {'max_iterations'} if isinstance(None,type(None)) else integers 
    try:
        out = wb.sheets[name].range('A1').options(dict,expand='table').value
        out2 = {k : int(v) if k in integers_ else v for k,v in out.items()}
    except:
        out2={}
    return out2
         



    
if __name__ == '__main__':
    testxls=Path('exceltest\lcrberegning2.xlsx')
    mmodel,para = getexcelmodel(testxls)
    eq=mmodel.equations
    mmodel.draw('_LCR_C62',up=10,down=1,HR=0,pdf=1)  # The LCR 
    mmodel.draw('_LCR_C25',up=4,down=1)  # liquid assets
    mmodel.draw('_LCR_C10',up=2)  # Leel 1 covered bonds 
    
    c,r = findcoordinates(testxls)
    xx = findequations(testxls)
